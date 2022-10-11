import pandas as pd
import xlwt
import openpyxl
from openpyxl import Workbook,load_workbook  # 引入头部文件
import re,nltk
import numpy as np

def is_critical_wordnet(word):
    function_list = ['CC', 'IN', 'LS', 'TO', 'POS', 'RP', 'SYM', 'UH']
    is_entity_critical_word = 1
    word_list = list(word)
    pos_tags = nltk.pos_tag(word_list)[0][1]
    if pos_tags in function_list:
        is_entity_critical_word = 0
    # number_of_sense_in_wordnet = len(wordnet.synsets(word))
    return is_entity_critical_word#,number_of_sense_in_wordnet

def get_character_feature(word):
    # number_of_characters=len(word)
    # word_length可以代替
    if word.istitle()==True:
        start_with_capital_letter=1
    else:
        start_with_capital_letter=0

    have_alphanumeric_letters=0
    pattern = re.compile('[0-9]+')
    for v in word:
        match = pattern.findall(v)
        if match:
            have_alphanumeric_letters=1
            break
    if word.isupper()==True:
        capital_letters_only=1
    else:
        capital_letters_only =0
    return start_with_capital_letter,have_alphanumeric_letters,capital_letters_only

wb = Workbook()
ws = wb.active
ws.title = 'data'

EA_avg = openpyxl.load_workbook('EA_avg.xlsx')
EA_avg = EA_avg.get_sheet_by_name('Sheet1')
A=EA_avg['A']
B=EA_avg['B']
C=EA_avg['C']
D=EA_avg['D']
E=EA_avg['E']
F=EA_avg['F']
G=EA_avg['G']
H=EA_avg['H']
I=EA_avg['I']
J=EA_avg['J']
K=EA_avg['K']
L=EA_avg['L']

name_list=['sentence_id', 'sentence', 'word_id', 'word',
                       'start_with_capital_letter', 'have_alphanumeric_letters', 'capital_letters_only',
                       'number_of_characters', 'is_entity_critical_word',
                       'number_of_dominated_nodes','complexity_score',
                       'max_dependency_distance','number_of_senses_in_wordnet',
                       'avg_word_first_fixation',  'avg_word_total_reading_time']

for j in range(len(name_list)):
    ws.cell(1, (j+1)).value=name_list[j]

for i in range(1,len(A)):
    print(i/(len(A)))
    word=D[i].value
    start_with_capital_letter, have_alphanumeric_letters, capital_letters_only=get_character_feature(word)
    is_entity_critical_word=is_critical_wordnet(word)
    title_list=[A[i].value,B[i].value,C[i].value,D[i].value,start_with_capital_letter, have_alphanumeric_letters, capital_letters_only,E[i].value,F[i].value,G[i].value,H[i].value,is_entity_critical_word,J[i].value,K[i].value,L[i].value]
    for j in range(len(title_list)):
        ws.cell((i+1), (j + 1)).value = title_list[j]

wb.save('EA.xlsx')